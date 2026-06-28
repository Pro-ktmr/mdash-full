#!/usr/bin/env python3
"""Summarize workbook layouts and keyword locations for extractor development."""

from __future__ import annotations

import argparse
import collections
import json
from pathlib import Path

from openpyxl import load_workbook


KEYWORDS = (
    "修了要件",
    "修了認定",
    "構成科目",
    "科目一覧",
    "授業科目",
    "科目名",
    "単位数",
)


def normalize(value: object) -> str:
    if value is None:
        return ""
    return " ".join(str(value).replace("\u3000", " ").split())


def analyze_file(path: Path) -> dict[str, object]:
    wb = load_workbook(path, read_only=False, data_only=False)
    sheets: list[dict[str, object]] = []
    for ws in wb.worksheets:
        hits: list[dict[str, str]] = []
        nonempty = 0
        for row in ws.iter_rows():
            for cell in row:
                text = normalize(cell.value)
                if not text:
                    continue
                nonempty += 1
                matched = [keyword for keyword in KEYWORDS if keyword in text]
                if matched:
                    hits.append(
                        {"cell": cell.coordinate, "text": text[:160], "keywords": matched}
                    )
        sheets.append(
            {
                "name": ws.title,
                "max_row": ws.max_row,
                "max_column": ws.max_column,
                "nonempty": nonempty,
                "hits": hits,
            }
        )
    wb.close()
    return {"file": str(path), "sheets": sheets}


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("root", type=Path, nargs="?", default=Path("literacy_application"))
    parser.add_argument("--sample", type=int, default=3)
    parser.add_argument("--all", action="store_true", help="Analyze every workbook")
    args = parser.parse_args()

    summary: dict[str, object] = {}
    for year_dir in sorted(p for p in args.root.iterdir() if p.is_dir()):
        files = sorted(year_dir.glob("*.xlsx"))
        selected = files if args.all else files[: args.sample]
        sheet_names: collections.Counter[str] = collections.Counter()
        records = []
        for path in selected:
            record = analyze_file(path)
            records.append(record)
            sheet_names.update(sheet["name"] for sheet in record["sheets"])
        summary[year_dir.name] = {
            "file_count": len(files),
            "analyzed_count": len(selected),
            "sheet_names": dict(sheet_names),
            "files": records,
        }

    print(json.dumps(summary, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
