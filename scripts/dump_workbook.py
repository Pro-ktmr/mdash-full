#!/usr/bin/env python3
"""Print non-empty cells from a workbook sheet for layout diagnosis."""

from __future__ import annotations

import argparse
from pathlib import Path

from openpyxl import load_workbook


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("workbook", type=Path)
    parser.add_argument("--sheet-contains", default="構成")
    parser.add_argument("--max-row", type=int, default=120)
    args = parser.parse_args()

    wb = load_workbook(args.workbook, read_only=False, data_only=False)
    sheets = [ws for ws in wb.worksheets if args.sheet_contains in ws.title]
    if not sheets:
        sheets = list(wb.worksheets[:1])
    for ws in sheets:
        print(f"## {ws.title} rows={ws.max_row} cols={ws.max_column}")
        print("MERGES", ", ".join(str(item) for item in ws.merged_cells.ranges))
        for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, args.max_row)):
            parts = []
            for cell in row:
                if cell.value is None or str(cell.value).strip() == "":
                    continue
                text = " ".join(str(cell.value).replace("\n", " ").split())
                parts.append(f"{cell.coordinate}={text[:220]}")
            if parts:
                print(" | ".join(parts))
    wb.close()
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
