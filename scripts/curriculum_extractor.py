#!/usr/bin/env python3
"""Year-aware curriculum extraction for MDASH literacy application workbooks."""

from __future__ import annotations

import argparse
import json
import math
import re
import sys
import unicodedata
from collections import Counter, defaultdict
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Iterable, Sequence

from openpyxl import load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.worksheet.worksheet import Worksheet


SCHEMA_VERSION = "1.0.0"
SUPPORTED_YEARS = ("r3", "r4", "r5", "r6", "r7", "r7_pre")
SECTION_PREFIX_RE = re.compile(r"^[①②③④⑤⑥⑦⑧⑨⑩⑪⑫⑬⑭⑮⑯⑰⑱⑲⑳]")
MODEL_CODE_RE = re.compile(r"^[1-4][\-‐‑‒–—―−][1-9]$")
MODEL_CODE_IN_TEXT_RE = re.compile(r"(?<!\d)([1-4])[\-‐‑‒–—―−]([1-9])(?!\d)")
NUMBER_RE = re.compile(r"[-+]?\d+(?:\.\d+)?")


@dataclass(frozen=True)
class YearConfig:
    year: str
    strategy: str
    main_max_col: int
    requirement_header: str


YEAR_CONFIGS: dict[str, YearConfig] = {
    "r3": YearConfig("r3", "r3_indexed", 10, "具体的な修了要件"),
    "r4": YearConfig("r4", "repeated_tables", 26, "具体的な修了要件"),
    "r5": YearConfig("r5", "repeated_tables", 19, "③修了要件"),
    "r6": YearConfig("r6", "repeated_tables", 21, "③修了要件"),
    "r7": YearConfig("r7", "repeated_tables", 21, "③修了要件"),
    "r7_pre": YearConfig("r7_pre", "direct_categories", 26, "④修了要件"),
}


def normalize_text(value: object) -> str:
    """Return human-readable text with Unicode and whitespace normalized."""
    if value is None:
        return ""
    if isinstance(value, bool):
        return "true" if value else "false"
    if isinstance(value, float) and math.isfinite(value) and value.is_integer():
        return str(int(value))
    text = unicodedata.normalize("NFKC", str(value))
    return " ".join(text.replace("\u3000", " ").split())


def compact_text(value: object) -> str:
    return re.sub(r"\s+", "", normalize_text(value))


def is_section_heading(value: object) -> bool:
    """Recognize numbered form sections without losing circled digits to NFKC."""
    if value is None:
        return False
    raw = re.sub(r"\s+", "", str(value).replace("\u3000", " "))
    if SECTION_PREFIX_RE.match(raw):
        return True
    compact = compact_text(value)
    return bool(
        re.match(
            r"^(?:[1-9]|1[0-9])(?:教育|対象|プログラム|具体|修了|現在|「|選択|授業|学修)",
            compact,
        )
    )


def canonical_course_name(value: object) -> str:
    text = compact_text(value).casefold()
    return text.replace("・", "·")


def normalize_model_code(value: object) -> str | None:
    text = normalize_text(value)
    match = MODEL_CODE_IN_TEXT_RE.search(text)
    return f"{match.group(1)}-{match.group(2)}" if match else None


def is_formula(value: object) -> bool:
    return isinstance(value, str) and value.startswith("=")


def json_number(value: object) -> int | float | None:
    if isinstance(value, bool):
        return None
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        if not math.isfinite(value):
            return None
        return int(value) if value.is_integer() else value
    text = normalize_text(value).replace(",", "")
    match = NUMBER_RE.search(text)
    if not match:
        return None
    number = float(match.group())
    return int(number) if number.is_integer() else number


def iter_cells(
    ws: Worksheet,
    *,
    min_row: int = 1,
    max_row: int | None = None,
    min_col: int = 1,
    max_col: int | None = None,
) -> Iterable[Cell]:
    row_limit = min(ws.max_row, max_row or ws.max_row)
    col_limit = min(ws.max_column, max_col or ws.max_column)
    for row in ws.iter_rows(
        min_row=min_row,
        max_row=max(min_row, row_limit),
        min_col=min_col,
        max_col=max(min_col, col_limit),
    ):
        yield from row


def cell_text(cell: Cell) -> str:
    return normalize_text(cell.value)


def is_target_sheet(ws: Worksheet) -> bool:
    title = compact_text(ws.title)
    if "構成する授業科目" in title:
        return True
    for cell in iter_cells(ws, max_row=12, max_col=min(ws.max_column, 30)):
        text = compact_text(cell.value)
        if "構成する授業科目について" in text:
            return True
    return False


def find_first_cell(
    ws: Worksheet,
    predicate,
    *,
    max_row: int = 250,
    max_col: int | None = None,
) -> Cell | None:
    for cell in iter_cells(ws, max_row=max_row, max_col=max_col):
        if predicate(cell_text(cell)):
            return cell
    return None


def next_nonempty_right(
    ws: Worksheet,
    cell: Cell,
    *,
    max_col: int,
    stop_at_section: bool = False,
) -> Cell | None:
    for col in range(cell.column + 1, min(ws.max_column, max_col) + 1):
        candidate = ws.cell(cell.row, col)
        text = cell_text(candidate)
        if not text or is_formula(candidate.value):
            continue
        if stop_at_section and is_section_heading(candidate.value):
            break
        return candidate
    return None


def find_labeled_value(
    worksheets: Sequence[Worksheet],
    labels: Sequence[str],
    *,
    max_col: int = 30,
) -> tuple[str | None, str | None, str | None]:
    label_keys = tuple(compact_text(label) for label in labels)
    for ws in worksheets:
        for cell in iter_cells(ws, max_row=120, max_col=max_col):
            current = compact_text(cell.value)
            if not current or not any(label in current for label in label_keys):
                continue
            right = next_nonempty_right(ws, cell, max_col=max_col, stop_at_section=True)
            if right and not any(label in compact_text(right.value) for label in label_keys):
                return cell_text(right), ws.title, right.coordinate
            # Older forms sometimes place the filled value in a merged block below.
            for row in range(cell.row + 1, min(ws.max_row, cell.row + 4) + 1):
                for col in range(cell.column, min(ws.max_column, max_col) + 1):
                    candidate = ws.cell(row, col)
                    text = cell_text(candidate)
                    if not text or is_formula(candidate.value):
                        continue
                    if is_section_heading(candidate.value):
                        break
                    return text, ws.title, candidate.coordinate
    return None, None, None


def find_section_boundary(ws: Worksheet, start_row: int, max_col: int) -> int:
    for row in range(start_row + 1, min(ws.max_row, 500) + 1):
        # Form section headings are anchored in column A. Restricting this check
        # avoids treating numbered requirement prose such as "① ... ② ..." in
        # a merged answer cell as the start of a new form section.
        if is_section_heading(ws.cell(row, 1).value):
            return row
    return min(ws.max_row, 500) + 1


def find_requirement_header(ws: Worksheet, config: YearConfig) -> Cell | None:
    target = compact_text(config.requirement_header)
    exact_matches: list[Cell] = []
    loose_matches: list[Cell] = []
    for cell in iter_cells(ws, max_row=80, max_col=config.main_max_col):
        text = compact_text(cell.value)
        if not text:
            continue
        if text == target:
            exact_matches.append(cell)
        elif target in text and "教育プログラムの修了要件" not in text:
            loose_matches.append(cell)
    if exact_matches:
        return exact_matches[0]
    return loose_matches[0] if loose_matches else None


def is_requirement_noise(text: str) -> bool:
    compact = compact_text(text)
    if not compact:
        return True
    if compact in {
        "科目",
        "単位",
        "【プルダウンリスト】",
        "教育プログラムの修了要件",
        "必要最低科目数",
        "必要最低単位数",
        "必要最低科目数・単位数",
        "履修必須の有無",
        "プログラム履修必須の有無",
    }:
        return True
    if re.fullmatch(r"\d+教育プログラムの修了要件", compact):
        return True
    if compact.startswith("学部・学科によって、修了要件") and len(compact) < 80:
        return True
    if compact.startswith("令和") and "履修" in compact and "必須" in compact:
        return True
    return False


def extract_completion_requirement(
    ws: Worksheet, config: YearConfig
) -> tuple[str | None, list[str]]:
    header = find_requirement_header(ws, config)
    if not header:
        return None, []
    boundary = find_section_boundary(ws, header.row, config.main_max_col)
    found: list[tuple[str, str]] = []

    # The pre-application form puts the answer to the right of the heading.
    for col in range(header.column + 1, config.main_max_col + 1):
        candidate = ws.cell(header.row, col)
        text = cell_text(candidate)
        if text and not is_formula(candidate.value) and not is_requirement_noise(text):
            found.append((text, candidate.coordinate))

    for row in range(header.row + 1, boundary):
        for col in range(1, config.main_max_col + 1):
            candidate = ws.cell(row, col)
            text = cell_text(candidate)
            if (
                not text
                or is_formula(candidate.value)
                or (
                    isinstance(candidate.value, (int, float))
                    and not isinstance(candidate.value, bool)
                )
                or is_requirement_noise(text)
            ):
                continue
            found.append((text, candidate.coordinate))

    unique: list[tuple[str, str]] = []
    seen = set()
    for text, coordinate in found:
        key = compact_text(text)
        if key in seen:
            continue
        seen.add(key)
        unique.append((text, coordinate))
    if not unique:
        return None, []
    return "\n".join(text for text, _ in unique), [coordinate for _, coordinate in unique]


def extract_requirement_scope(ws: Worksheet, max_col: int) -> dict[str, Any]:
    for cell in iter_cells(ws, max_row=60, max_col=max_col):
        text = cell_text(cell)
        compact = compact_text(text)
        if "学部・学科によって、修了要件は相違" not in compact:
            continue
        if "場合は" in compact or "複製" in compact or "プルダウン" in compact:
            continue
        same = None
        if "相違しない" in compact:
            same = True
        elif "相違する" in compact:
            same = False
        return {"raw": text, "same_across_departments": same, "source_cell": cell.coordinate}
    return {"raw": None, "same_across_departments": None, "source_cell": None}


def extract_target_departments(ws: Worksheet, max_col: int) -> tuple[list[str], list[str]]:
    header = find_first_cell(
        ws,
        lambda text: "対象となる学部・学科名称" in compact_text(text),
        max_row=50,
        max_col=max_col,
    )
    if not header:
        return [], []
    boundary = find_section_boundary(ws, header.row, max_col)
    values: list[tuple[str, str]] = []

    # Same-row answers exist in the redesigned form. Stop at a competing section heading.
    for col in range(header.column + 1, max_col + 1):
        candidate = ws.cell(header.row, col)
        text = cell_text(candidate)
        if not text:
            continue
        if is_section_heading(candidate.value):
            break
        values.append((text, candidate.coordinate))

    for row in range(header.row + 1, boundary):
        for col in range(1, max_col + 1):
            candidate = ws.cell(row, col)
            text = cell_text(candidate)
            compact = compact_text(text)
            if not text or is_formula(candidate.value):
                continue
            if "教育プログラムの修了要件" in compact or "学部・学科によって" in compact:
                continue
            values.append((text, candidate.coordinate))

    result: list[str] = []
    cells: list[str] = []
    seen = set()
    for text, coordinate in values:
        key = compact_text(text)
        if not key or key in seen:
            continue
        seen.add(key)
        result.append(text)
        cells.append(coordinate)
    return result, cells


def extract_program_required_status(ws: Worksheet, max_col: int) -> tuple[str | None, str | None]:
    label = find_first_cell(
        ws,
        lambda text: "履修必須の有無" in compact_text(text),
        max_row=80,
        max_col=max_col,
    )
    if not label:
        return None, None
    value = next_nonempty_right(ws, label, max_col=max_col, stop_at_section=True)
    if not value:
        return None, None
    return cell_text(value), value.coordinate


def nearest_number_left(ws: Worksheet, row: int, col: int, min_col: int) -> int | float | None:
    for candidate_col in range(col - 1, min_col - 1, -1):
        value = json_number(ws.cell(row, candidate_col).value)
        if value is not None:
            return value
    return None


def infer_minimum_from_text(requirement: str | None, unit: str) -> int | float | None:
    if not requirement:
        return None
    normalized = unicodedata.normalize("NFKC", requirement).replace(" ", "")
    patterns = [
        rf"合計([0-9]+(?:\.[0-9]+)?){unit}以上",
        rf"最低([0-9]+(?:\.[0-9]+)?){unit}",
        rf"([0-9]+(?:\.[0-9]+)?){unit}以上(?:を)?(?:取得|修得)",
        rf"([0-9]+(?:\.[0-9]+)?){unit}(?:を)?(?:取得|修得)",
    ]
    for pattern in patterns:
        matches = re.findall(pattern, normalized)
        if matches:
            value = float(matches[-1])
            return int(value) if value.is_integer() else value
    return None


def extract_minimums(
    ws: Worksheet, max_col: int, requirement: str | None
) -> tuple[int | float | None, int | float | None, str | None]:
    label = find_first_cell(
        ws,
        lambda text: "必要最低" in compact_text(text),
        max_row=100,
        max_col=max_col,
    )
    courses = None
    credits = None
    source = label.coordinate if label else None
    if label:
        for row in range(label.row, min(ws.max_row, label.row + 2) + 1):
            for col in range(label.column + 1, max_col + 1):
                text = compact_text(ws.cell(row, col).value)
                if text == "科目":
                    courses = nearest_number_left(ws, row, col, label.column + 1)
                elif text == "単位":
                    credits = nearest_number_left(ws, row, col, label.column + 1)
    courses = courses if courses is not None else infer_minimum_from_text(requirement, "科目")
    credits = credits if credits is not None else infer_minimum_from_text(requirement, "単位")
    return courses, credits, source


def course_category_from_r3(name: str) -> tuple[str, str | None, bool | None]:
    normalized = normalize_text(name)
    match = re.search(r"\s*[\(（]\s*(必修科目|基幹科目|連携科目)\s*[\)）]\s*$", normalized)
    if not match:
        return normalized, None, None
    category = match.group(1)
    clean = normalized[: match.start()].rstrip()
    return clean or normalized, category, category == "必修科目"


def finalize_course_groups(
    groups: dict[str, list[dict[str, Any]]], warnings: list[str]
) -> list[dict[str, Any]]:
    courses: list[dict[str, Any]] = []
    for occurrences in groups.values():
        names = list(dict.fromkeys(item["name"] for item in occurrences))
        credit_values = [item.get("credits") for item in occurrences if item.get("credits") is not None]
        unique_credits = list(dict.fromkeys(credit_values))
        if len(unique_credits) > 1:
            warnings.append(f"conflicting_credits:{names[0]}:{unique_credits}")
        required_values = [item.get("is_required") for item in occurrences if item.get("is_required") is not None]
        requirement_types = list(
            dict.fromkeys(item["requirement_type"] for item in occurrences if item.get("requirement_type"))
        )
        categories = list(dict.fromkeys(item["category"] for item in occurrences if item.get("category")))
        codes = sorted(
            {
                code
                for item in occurrences
                for code in item.get("model_curriculum_codes", [])
                if code
            },
            key=lambda code: tuple(int(part) for part in code.split("-")),
        )
        covered_sections = list(
            dict.fromkeys(item["section"] for item in occurrences if item.get("section"))
        )
        course = {
            "name": names[0],
            "name_variants": names[1:],
            "credits": unique_credits[0] if unique_credits else None,
            "is_required": True if True in required_values else (False if False in required_values else None),
            "requirement_types": requirement_types,
            "categories": categories,
            "model_curriculum_codes": codes,
            "covered_sections": covered_sections,
            "source_occurrences": [
                {
                    key: value
                    for key, value in item.items()
                    if key
                    in {
                        "source_cell",
                        "section",
                        "credits",
                        "required_raw",
                        "opening_status",
                        "optional_item",
                        "model_curriculum_codes",
                        "listed_number",
                    }
                    and value not in (None, "", [])
                }
                for item in occurrences
            ],
        }
        courses.append(course)
    return courses


def extract_r3_courses(ws: Worksheet, config: YearConfig) -> tuple[list[dict[str, Any]], list[str]]:
    warnings: list[str] = []
    header_rows: dict[int, list[Cell]] = defaultdict(list)
    for cell in iter_cells(ws, max_row=100, max_col=config.main_max_col):
        if compact_text(cell.value) in {"授業科目名称", "授業科目名", "科目名"}:
            header_rows[cell.row].append(cell)
    if not header_rows:
        return [], ["course_header_not_found"]
    header_row = max(header_rows, key=lambda row: (len(header_rows[row]), row))

    groups: dict[str, list[dict[str, Any]]] = defaultdict(list)
    table_headers = sorted(header_rows[header_row], key=lambda cell: cell.column)
    for header_index, header in enumerate(table_headers):
        index_col = header.column
        block_end_col = (
            table_headers[header_index + 1].column - 1
            if header_index + 1 < len(table_headers)
            else config.main_max_col
        )
        for row in range(header_row + 1, min(ws.max_row, 500) + 1):
            number = json_number(ws.cell(row, index_col).value)
            candidates = [
                ws.cell(row, col)
                for col in range(index_col + 1, block_end_col + 1)
                if cell_text(ws.cell(row, col)) and not is_formula(ws.cell(row, col).value)
            ]
            if not candidates:
                continue
            # Some r3 workbooks insert a merged learning-outcome category before
            # the actual subject. The rightmost filled cell in the table block is
            # the subject name; standard forms still resolve to column B/G.
            name_cell = candidates[-1]
            raw_name = cell_text(name_cell)
            if number is None or number < 1 or not raw_name or is_formula(name_cell.value):
                continue
            name, category, is_required = course_category_from_r3(raw_name)
            key = canonical_course_name(name)
            if not key:
                continue
            groups[key].append(
                {
                    "name": name,
                    "credits": None,
                    "is_required": is_required,
                    "requirement_type": "required" if is_required else ("listed" if category else None),
                    "category": category,
                    "model_curriculum_codes": [],
                    "section": "授業科目名称",
                    "source_cell": name_cell.coordinate,
                    "listed_number": number,
                }
            )
    return finalize_course_groups(groups, warnings), warnings


def section_rows(ws: Worksheet, max_col: int) -> list[tuple[int, str]]:
    found: dict[int, str] = {}
    for cell in iter_cells(ws, max_row=250, max_col=1):
        text = cell_text(cell)
        if is_section_heading(cell.value):
            found.setdefault(cell.row, text)
    return sorted(found.items())


def nearest_section(sections: list[tuple[int, str]], row: int) -> str | None:
    candidates = [text for section_row, text in sections if section_row <= row]
    return candidates[-1] if candidates else None


def next_section_row(sections: list[tuple[int, str]], row: int, fallback: int) -> int:
    return next((section_row for section_row, _ in sections if section_row > row), fallback)


def marked(value: object) -> bool:
    text = compact_text(value)
    return bool(text and text not in {"-", "―", "0", "×"})


def parse_required_marker(value: object) -> bool | None:
    text = compact_text(value)
    if not text:
        return None
    if text in {"○", "〇", "必須", "必修", "1", "true", "yes"}:
        return True
    if text in {"×", "選択", "任意", "0", "false", "no"}:
        return False
    return None


def extract_repeated_table_courses(
    ws: Worksheet, config: YearConfig
) -> tuple[list[dict[str, Any]], list[str]]:
    warnings: list[str] = []
    valid_headers = {"授業科目", "授業科目名", "科目名"}
    headers_by_row: dict[int, list[Cell]] = defaultdict(list)
    for cell in iter_cells(ws, max_row=180, max_col=config.main_max_col):
        if compact_text(cell.value) in valid_headers:
            headers_by_row[cell.row].append(cell)
    if not headers_by_row:
        return [], ["course_header_not_found"]

    sections = section_rows(ws, config.main_max_col)
    groups: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for header_row, row_headers in sorted(headers_by_row.items()):
        row_headers.sort(key=lambda cell: cell.column)
        data_end = next_section_row(
            sections, header_row, min(ws.max_row, header_row + 12) + 1
        )
        section = nearest_section(sections, header_row)
        for index, header in enumerate(row_headers):
            block_end_col = (
                row_headers[index + 1].column - 1
                if index + 1 < len(row_headers)
                else config.main_max_col
            )
            header_map: dict[str, int] = {}
            model_columns: dict[int, str] = {}
            for col in range(header.column + 1, block_end_col + 1):
                text = cell_text(ws.cell(header_row, col))
                compact = compact_text(text)
                if compact in {"単位数", "必須", "必修", "開講状況", "選択項目"}:
                    header_map.setdefault(compact, col)
                if MODEL_CODE_RE.fullmatch(compact):
                    code = normalize_model_code(compact)
                    if code:
                        model_columns[col] = code

            credits_col = header_map.get("単位数")
            required_col = header_map.get("必須") or header_map.get("必修")
            opening_col = header_map.get("開講状況")
            optional_col = header_map.get("選択項目")

            for row in range(header_row + 1, data_end):
                name_cell = ws.cell(row, header.column)
                raw_name = cell_text(name_cell)
                if (
                    not raw_name
                    or is_formula(name_cell.value)
                    or compact_text(raw_name).startswith("※")
                    or compact_text(raw_name) in valid_headers
                    or is_section_heading(name_cell.value)
                ):
                    continue
                credits = json_number(ws.cell(row, credits_col).value) if credits_col else None
                required_raw = cell_text(ws.cell(row, required_col)) if required_col else None
                optional_item = cell_text(ws.cell(row, optional_col)) if optional_col else None
                model_codes = [
                    code for col, code in model_columns.items() if marked(ws.cell(row, col).value)
                ]
                optional_code = normalize_model_code(optional_item)
                if optional_code and optional_code not in model_codes:
                    model_codes.append(optional_code)
                is_required = parse_required_marker(required_raw)
                requirement_type = None
                if is_required is True:
                    requirement_type = "required"
                elif optional_col is not None or (section and "選択" in section):
                    requirement_type = "optional"
                    if is_required is None:
                        is_required = False
                elif is_required is False:
                    requirement_type = "optional"

                key = canonical_course_name(raw_name)
                if not key:
                    continue
                groups[key].append(
                    {
                        "name": raw_name,
                        "credits": credits,
                        "is_required": is_required,
                        "requirement_type": requirement_type,
                        "category": None,
                        "model_curriculum_codes": model_codes,
                        "section": section,
                        "source_cell": name_cell.coordinate,
                        "required_raw": required_raw,
                        "opening_status": (
                            cell_text(ws.cell(row, opening_col)) if opening_col else None
                        ),
                        "optional_item": optional_item,
                    }
                )
    return finalize_course_groups(groups, warnings), warnings


def direct_category_ranges(
    ws: Worksheet, start_row: int, end_row: int
) -> list[tuple[int, int, str, str]]:
    ranges: list[tuple[int, int, str, str]] = []
    for merged in ws.merged_cells.ranges:
        if merged.max_row < start_row or merged.min_row > end_row or merged.min_col > 3:
            continue
        text = cell_text(ws.cell(merged.min_row, merged.min_col))
        compact = compact_text(text)
        if re.match(r"^\([123]\)", compact):
            ranges.append((merged.min_row, merged.max_row, text, ws.cell(merged.min_row, merged.min_col).coordinate))
    for row in range(start_row, end_row + 1):
        for col in range(1, 4):
            cell = ws.cell(row, col)
            text = cell_text(cell)
            if re.match(r"^\([123]\)", compact_text(text)) and not any(
                start <= row <= end for start, end, _, _ in ranges
            ):
                ranges.append((row, row, text, cell.coordinate))
    return sorted(ranges)


def category_type(category: str | None) -> tuple[str | None, bool | None]:
    compact = compact_text(category)
    if compact.startswith("(1)") or "必須科目" in compact:
        return "required", True
    if compact.startswith("(2)") or "選択必須科目" in compact:
        return "required_elective", None
    if compact.startswith("(3)") or "選択科目" in compact:
        return "elective", False
    return None, None


def extract_direct_category_courses(
    ws: Worksheet, config: YearConfig
) -> tuple[list[dict[str, Any]], list[str]]:
    warnings: list[str] = []
    header = find_first_cell(
        ws,
        lambda text: compact_text(text) in {"授業科目", "授業科目名", "科目名"},
        max_row=100,
        max_col=config.main_max_col,
    )
    if not header:
        return [], ["course_header_not_found"]
    sections = section_rows(ws, config.main_max_col)
    end_row = next_section_row(sections, header.row, min(ws.max_row, header.row + 40) + 1) - 1
    credits_col = None
    model_columns: dict[int, str] = {}
    for col in range(header.column + 1, config.main_max_col + 1):
        text = cell_text(ws.cell(header.row, col))
        if compact_text(text) == "単位数" and credits_col is None:
            credits_col = col
        code = normalize_model_code(text)
        if code and MODEL_CODE_RE.fullmatch(compact_text(text)):
            model_columns[col] = code

    categories = direct_category_ranges(ws, header.row + 1, end_row)
    groups: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in range(header.row + 1, end_row + 1):
        name_cell = ws.cell(row, header.column)
        raw_name = cell_text(name_cell)
        if not raw_name or is_formula(name_cell.value):
            continue
        category_record = next(
            ((text, source) for start, end, text, source in categories if start <= row <= end),
            (None, None),
        )
        category, _category_source = category_record
        requirement_type, is_required = category_type(category)
        model_codes = [
            code for col, code in model_columns.items() if marked(ws.cell(row, col).value)
        ]
        key = canonical_course_name(raw_name)
        if not key:
            continue
        groups[key].append(
            {
                "name": raw_name,
                "credits": json_number(ws.cell(row, credits_col).value) if credits_col else None,
                "is_required": is_required,
                "requirement_type": requirement_type,
                "category": category,
                "model_curriculum_codes": model_codes,
                "section": "プログラム構成科目",
                "source_cell": name_cell.coordinate,
            }
        )
    return finalize_course_groups(groups, warnings), warnings


def extract_courses(
    ws: Worksheet, config: YearConfig
) -> tuple[list[dict[str, Any]], list[str]]:
    if config.strategy == "r3_indexed":
        return extract_r3_courses(ws, config)
    if config.strategy == "repeated_tables":
        return extract_repeated_table_courses(ws, config)
    if config.strategy == "direct_categories":
        return extract_direct_category_courses(ws, config)
    raise ValueError(f"Unknown strategy: {config.strategy}")


def extract_program(
    ws: Worksheet,
    all_sheets: Sequence[Worksheet],
    config: YearConfig,
) -> dict[str, Any]:
    warnings: list[str] = []
    requirement, requirement_cells = extract_completion_requirement(ws, config)
    courses, course_warnings = extract_courses(ws, config)
    for course in courses:
        for occurrence in course["source_occurrences"]:
            occurrence["sheet_name"] = ws.title
    warnings.extend(course_warnings)
    departments, department_cells = extract_target_departments(ws, config.main_max_col)
    minimum_courses, minimum_credits, minimum_source = extract_minimums(
        ws, config.main_max_col, requirement
    )
    required_status, required_status_cell = extract_program_required_status(
        ws, config.main_max_col
    )
    program_name, program_sheet, program_cell = find_labeled_value(
        [ws, *[sheet for sheet in all_sheets if sheet is not ws]],
        ("プログラム名", "申請するプログラム又は授業科目名称"),
        max_col=30,
    )

    if not requirement:
        warnings.append("completion_requirement_not_found")
    if not courses:
        warnings.append("courses_not_found")
    status = "success" if requirement and courses else "partial"
    return {
        "sheet_name": ws.title,
        "source_sheets": [ws.title],
        "program_name": program_name,
        "target_departments": departments,
        "completion_requirement": {
            "text": requirement,
            "minimum_courses": minimum_courses,
            "minimum_credits": minimum_credits,
            "scope": extract_requirement_scope(ws, config.main_max_col),
            "program_required_status": required_status,
        },
        "courses": courses,
        "course_count": len(courses),
        "extraction": {
            "status": status,
            "strategy": config.strategy,
            "warnings": list(dict.fromkeys(warnings)),
            "source": {
                "requirement_cells": requirement_cells,
                "minimums_cell": minimum_source,
                "department_cells": department_cells,
                "program_name": (
                    {"sheet": program_sheet, "cell": program_cell}
                    if program_sheet and program_cell
                    else None
                ),
                "program_required_status_cell": required_status_cell,
            },
        },
    }


def extend_unique(target: list[Any], values: Iterable[Any]) -> None:
    for value in values:
        if value not in target:
            target.append(value)


def merge_course_records(
    target_courses: list[dict[str, Any]],
    incoming_courses: Iterable[dict[str, Any]],
    warnings: list[str],
) -> None:
    by_name = {canonical_course_name(course["name"]): course for course in target_courses}
    for incoming in incoming_courses:
        key = canonical_course_name(incoming["name"])
        if key not in by_name:
            target_courses.append(incoming)
            by_name[key] = incoming
            continue
        target = by_name[key]
        if target["credits"] is None:
            target["credits"] = incoming["credits"]
        elif incoming["credits"] is not None and target["credits"] != incoming["credits"]:
            warnings.append(
                f"conflicting_credits:{target['name']}:[{target['credits']}, {incoming['credits']}]"
            )
        if incoming["name"] != target["name"]:
            extend_unique(target["name_variants"], [incoming["name"]])
        extend_unique(target["name_variants"], incoming["name_variants"])
        extend_unique(target["requirement_types"], incoming["requirement_types"])
        extend_unique(target["categories"], incoming["categories"])
        extend_unique(target["model_curriculum_codes"], incoming["model_curriculum_codes"])
        extend_unique(target["covered_sections"], incoming["covered_sections"])
        target["source_occurrences"].extend(incoming["source_occurrences"])
        if incoming["is_required"] is True:
            target["is_required"] = True
        elif target["is_required"] is None and incoming["is_required"] is False:
            target["is_required"] = False


def merge_continuation_programs(programs: list[dict[str, Any]]) -> list[dict[str, Any]]:
    """Merge split course-list pages when one shared requirement applies to all."""
    complete = [program for program in programs if program["completion_requirement"]["text"]]
    continuation = [
        program
        for program in programs
        if not program["completion_requirement"]["text"] and program["courses"]
    ]
    if len(complete) != 1 or not continuation or len(complete) + len(continuation) != len(programs):
        return programs
    base = complete[0]
    if base["completion_requirement"]["scope"]["same_across_departments"] is not True:
        return programs

    notes = base["extraction"].setdefault("notes", [])
    for page in continuation:
        merge_course_records(base["courses"], page["courses"], base["extraction"]["warnings"])
        extend_unique(base["source_sheets"], page["source_sheets"])
        extend_unique(base["target_departments"], page["target_departments"])
        notes.append(f"merged_continuation_sheet:{page['sheet_name']}")
    base["course_count"] = len(base["courses"])
    base["extraction"]["warnings"] = list(dict.fromkeys(base["extraction"]["warnings"]))
    return [base]


def repair_requirements_from_cached_values(
    programs: list[dict[str, Any]],
    cached_sheets: dict[str, Worksheet],
    config: YearConfig,
) -> None:
    """Recover formula-backed requirement text from Excel's saved calculation cache."""
    for program in programs:
        if program["completion_requirement"]["text"]:
            continue
        cached_ws = cached_sheets.get(program["sheet_name"])
        if cached_ws is None:
            continue
        requirement, source_cells = extract_completion_requirement(cached_ws, config)
        if not requirement:
            continue
        minimum_courses, minimum_credits, minimum_source = extract_minimums(
            cached_ws, config.main_max_col, requirement
        )
        program["completion_requirement"]["text"] = requirement
        program["completion_requirement"]["minimum_courses"] = minimum_courses
        program["completion_requirement"]["minimum_credits"] = minimum_credits
        program["extraction"]["source"]["requirement_cells"] = source_cells
        program["extraction"]["source"]["minimums_cell"] = minimum_source
        program["extraction"]["warnings"] = [
            warning
            for warning in program["extraction"]["warnings"]
            if warning != "completion_requirement_not_found"
        ]
        program["extraction"].setdefault("notes", []).append(
            "completion_requirement_from_cached_formula_value"
        )
        if program["courses"]:
            program["extraction"]["status"] = "success"


def extract_workbook(path: Path, config: YearConfig) -> dict[str, Any]:
    try:
        workbook = load_workbook(path, read_only=False, data_only=False, keep_links=False)
    except Exception as exc:  # noqa: BLE001 - the error is part of the report
        return {
            "university_name": path.stem,
            "source_file": str(path.as_posix()),
            "programs": [],
            "extraction": {
                "status": "failed",
                "warnings": [f"workbook_open_error:{type(exc).__name__}:{exc}"],
            },
        }

    try:
        sheets = list(workbook.worksheets)
        targets = [ws for ws in sheets if is_target_sheet(ws)]
        institution, institution_sheet, institution_cell = find_labeled_value(
            targets + [ws for ws in sheets if ws not in targets],
            ("大学等名", "学校名"),
            max_col=30,
        )
        university_name = institution or path.stem
        if not targets:
            return {
                "university_name": university_name,
                "source_file": str(path.as_posix()),
                "programs": [],
                "extraction": {
                    "status": "failed",
                    "warnings": ["target_sheet_not_found"],
                    "source": {
                        "university_name": (
                            {"sheet": institution_sheet, "cell": institution_cell}
                            if institution_sheet and institution_cell
                            else "filename"
                        )
                    },
                },
            }
        programs = [extract_program(ws, sheets, config) for ws in targets]
        if any(not program["completion_requirement"]["text"] for program in programs):
            cached_workbook = None
            try:
                cached_workbook = load_workbook(
                    path, read_only=False, data_only=True, keep_links=False
                )
                repair_requirements_from_cached_values(
                    programs,
                    {sheet.title: sheet for sheet in cached_workbook.worksheets},
                    config,
                )
            finally:
                if cached_workbook is not None:
                    cached_workbook.close()
        programs = merge_continuation_programs(programs)
        program_statuses = [program["extraction"]["status"] for program in programs]
        status = "success" if all(item == "success" for item in program_statuses) else "partial"
        warnings = [
            f"{program['sheet_name']}:{warning}"
            for program in programs
            for warning in program["extraction"]["warnings"]
        ]
        return {
            "university_name": university_name,
            "source_file": str(path.as_posix()),
            "programs": programs,
            "extraction": {
                "status": status,
                "warnings": warnings,
                "source": {
                    "university_name": (
                        {"sheet": institution_sheet, "cell": institution_cell}
                        if institution_sheet and institution_cell
                        else "filename"
                    )
                },
            },
        }
    except Exception as exc:  # noqa: BLE001 - continue across all universities
        return {
            "university_name": path.stem,
            "source_file": str(path.as_posix()),
            "programs": [],
            "extraction": {
                "status": "failed",
                "warnings": [f"extraction_error:{type(exc).__name__}:{exc}"],
            },
        }
    finally:
        workbook.close()


def year_summary(universities: list[dict[str, Any]]) -> dict[str, Any]:
    counts = Counter(item["extraction"]["status"] for item in universities)
    return {
        "university_count": len(universities),
        "status_counts": {
            "success": counts.get("success", 0),
            "partial": counts.get("partial", 0),
            "failed": counts.get("failed", 0),
        },
        "program_count": sum(len(item["programs"]) for item in universities),
        "course_count": sum(
            program["course_count"]
            for item in universities
            for program in item["programs"]
        ),
    }


def extract_year(
    year: str,
    input_dir: Path,
    *,
    progress: bool = False,
) -> dict[str, Any]:
    if year not in YEAR_CONFIGS:
        raise ValueError(f"Unsupported year: {year}")
    config = YEAR_CONFIGS[year]
    files = sorted(input_dir.glob("*.xlsx"))
    universities: list[dict[str, Any]] = []
    for index, path in enumerate(files, start=1):
        universities.append(extract_workbook(path, config))
        if progress and (index % 20 == 0 or index == len(files)):
            print(f"[{year}] {index}/{len(files)}", file=sys.stderr, flush=True)
    return {
        "schema_version": SCHEMA_VERSION,
        "year": year,
        "input_directory": str(input_dir.as_posix()),
        "summary": year_summary(universities),
        "universities": universities,
    }


def write_json(path: Path, payload: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    temporary = path.with_suffix(path.suffix + ".tmp")
    temporary.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    temporary.replace(path)


def run_year_cli(year: str) -> int:
    parser = argparse.ArgumentParser(
        description=f"Extract completion requirements and courses from {year} workbooks."
    )
    parser.add_argument(
        "--input-dir",
        type=Path,
        default=Path("literacy_application") / year,
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=Path("outputs") / "curriculum" / f"{year}.json",
    )
    parser.add_argument("--progress", action="store_true")
    args = parser.parse_args()
    payload = extract_year(year, args.input_dir, progress=args.progress)
    write_json(args.output, payload)
    print(json.dumps(payload["summary"], ensure_ascii=False))
    return 0
